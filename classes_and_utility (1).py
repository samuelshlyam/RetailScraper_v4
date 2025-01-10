
# Import statements
from typing import Any
import requests, datetime
from bs4 import BeautifulSoup
import json
import csv
import re
import logging
import random
from urllib.parse import urlparse
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import subprocess
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import unittest
import io
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
class Product:
    def __init__(self, input_sku, brand):
        self.input_sku = input_sku
        self.brand = brand
        self.variations = []
        self.title = None
        self.images = []
        self.prices = []
        self.currency = None
        self.url = None
        self.description = None
        self.source_type = None  # New attribute to store the source type
        self.seller = None  # New attribute to store the seller information
        self.excel_row_number=None
    def add_variation(self, variation):
        self.variations.append(variation)

    def set_details(self, title, images, prices, currency, url, description, seller):
        self.title = title
        self.images = images
        self.prices = prices
        self.currency = currency
        self.url = url
        self.description = description
        self.seller = seller  

    def is_complete(self):
        return bool(self.prices and self.url)

class DataFetcher:
    @staticmethod
    def parse_google_results(html_content):
        soup = BeautifulSoup(html_content, 'html.parser')
        results = []
        for g in soup.find_all('div', class_='g'):
            links = g.find_all('a')
            if links and 'href' in links[0].attrs:  # check if 'href' attribute exists
                results.append(links[0]['href'])
        return results
    def extract_product_schema(self, html_content):
        product_schemas = []  # List to store all found product schemas

        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            schema_tags = soup.find_all('script', {'type': 'application/ld+json'})

            for tag in schema_tags:
                try:
                    data = json.loads(tag.text)
                    if data.get('@type') == 'Product':
                        # Log the raw product schema for debugging
                        logging.debug("Raw Product Schema: %s", json.dumps(data, indent=4))
                        product_schemas.append(data)
                except json.JSONDecodeError:
                    continue

            if not product_schemas:
                logging.warning("No Product schema found in the HTML content.")
                return None

            return product_schemas
        except Exception as e:
            logging.error(f"Error extracting product schemas from HTML: {e}")
            return None
class SKUManager:
    def __init__(self, brand_settings):
        self.brand_settings = brand_settings
        
    def generate_variations(self, input_sku, brand_rule):
        brand_variations = self.handle_brand_sku(input_sku, brand_rule)
        blind_variations = self.handle_sku(input_sku, brand_rule)
        total_variations = brand_variations + blind_variations
        #modified_variations = []  # Create a new list for the modified variations
        #for variation in total_variations:
        #    modified_variations.append(f"site:farfetch.com {variation}")
        ## Combine the original and modified variations
        #total_variations += modified_variations
        return total_variations

    def handle_brand_sku(self, sku, brand_rule):
        cleaned_sku = self.clean_sku(sku)
        sku_format = brand_rule['sku_format']
        base_parts = sku_format['base']
        base_separator = sku_format.get('base_separator', '')
        color_separator = sku_format.get('color_separator', '')
        complex=brand_rule.get('complex', False)

        article_length = int(base_parts['article'][0].split(',')[0])
        model_length = int(base_parts['model'][0].split(',')[0])
        color_length = int(sku_format['color_extension'][0].split(',')[0])

        article_part = cleaned_sku[:article_length]
        model_part = cleaned_sku[article_length:article_length + model_length]
        color_part = cleaned_sku[article_length + model_length:article_length + model_length + color_length]
        if complex:
            return self.complex(article_length,model_length,cleaned_sku, base_separator)
        # Order: Brand Format with Color, Brand Format without Color
        return [
        article_part + base_separator + model_part + color_separator + color_part,
        article_part + base_separator + model_part
    ]

    def handle_sku(self, sku, brand_rule):
        cleaned_sku = self.clean_sku(sku)
        sku_format = brand_rule['sku_format']
        base_separator = sku_format.get('base_separator', '')
        article_length = int(sku_format['base']['article'][0].split(',')[0])
        model_length = int(sku_format['base']['model'][0].split(',')[0])
        color_length = int(sku_format['color_extension'][0].split(',')[0])
        complex=brand_rule.get('complex', False)
        article_part = cleaned_sku[:article_length]
        model_part = cleaned_sku[article_length:article_length + model_length]
        color_part = cleaned_sku[article_length + model_length:article_length + model_length + color_length]
        if complex:
            return self.complex(article_length,model_length,cleaned_sku, base_separator)
        # Order: No space (Article Model Color), Space (Article Model Color), No space (Article Model), Space (Article Model)
        return [
            article_part + model_part + color_part, 
            article_part + ' ' + model_part + ' ' + color_part,
            article_part + model_part,
            article_part + ' ' + model_part
        ]

    def complex(self, article_length,model_length,cleaned_sku, base_separator):
        print(cleaned_sku)
        new_sku=SKUManager.remove_letters_from_end(cleaned_sku)
        
        article_part = new_sku[:article_length]
        model_part = new_sku[len(new_sku)-model_length:]
        return [
            article_part + base_separator + model_part ,
            model_part
        ]
        
    @staticmethod
    def remove_letters_from_end(s):
        non_letter_index = None
        for i in range(len(s) - 1, -1, -1):
            if not s[i].isalpha():
                non_letter_index = i
                break

        # If there are no non-letter characters, return an empty string
        if non_letter_index is None:
            return ""

        # Return the string up to the last non-letter character
        return s[:non_letter_index + 1]
    @staticmethod
    def clean_sku(sku):
        sku = str(sku)
        Logger.log(f"Cleaning SKU: {sku}")
        cleaned = re.sub(r'[^a-zA-Z0-9]', '', sku)
        Logger.log(f"Cleaned SKU: {cleaned}")
        return cleaned
    @staticmethod
    def listify_file(file_path):
        with open(file_path, 'r') as file:
            reader = csv.reader(file)
            data = list(reader)
        return data
class BrandSettings:
    def __init__(self, settings):
        self.settings = settings

    def get_rules_for_brand(self, brand_name):
        for rule in self.settings['brand_rules']:
            if str(brand_name).lower() in str(rule['names']).lower():
                return rule
        return None 
import threading   
class SearchEngine:
    def __init__(self, user_agents):
        self.user_agents = user_agents
        
    def create_brand_search_query(self, sku, brand_settings, iteration):
        if iteration <= 1:  # For the first two iterations, include brand and site operator
            domain = brand_settings.get("domain_hierarchy", [])[0] if brand_settings.get("domain_hierarchy") else ""
            query = f"site:{domain} \"{sku}\""
        else:  # For the rest, just use the SKU
            query = f"\"{sku}\""
        #else:
        #    query = f"site:farfetch.com\"{sku.strip('site:farfetch.com')}\""
        return f"https://www.google.com/search?q={query}"

    def choose_random_header(self):
        ua = random.choice(self.user_agents)
        #return ua.replace(";", "")
        return ua
 
    def filter_urls_by_brand_and_whitelist(self, urls, brand_settings, whitelisted_domains):
        brand_domains = [domain.replace('www.', '') for domain in brand_settings.get("domain_hierarchy", [])]
        whitelisted_domains = [domain.replace('www.', '') for domain in whitelisted_domains]
        approved_brand_urls = []
        approved_whitelist_urls = []
        #approved_farfetch_urls=[]
        approved_modesens_urls=[]
        if isinstance(urls, str):
            urls = urls.split(',')

        for url in urls:
            url = str(url).strip()
            if not url.startswith(('http://', 'https://')):
                url = 'http://' + url

            try:
                parsed_url = urlparse(url)
                domain = parsed_url.netloc
                if domain.startswith('www.'):
                    domain = domain[4:]
                Logger.log(f"Domain: {domain}")
                if domain in brand_domains:
                    approved_brand_urls.append([url, "brand"])
                elif domain in whitelisted_domains:
                    approved_whitelist_urls.append([url, "whitelist"])
                #elif "farfetch" in domain:
                #    approved_farfetch_urls.append([url, "farfetch"])
                elif 'modesens' in domain:
                    approved_modesens_urls.append([url, "modesens"])
            except Exception as e:
                Logger.log(f"Error parsing URL '{url}': {e}")
        #Combine brand URLs and whitelist URLs
        approved_urls=approved_brand_urls+approved_whitelist_urls+approved_modesens_urls#+approved_farfetch_urls
        return approved_urls

################################
    def filter_urls_by_currency(self, currency_items, urls):
        filtered_urls = []
        Logger.log(f'Filtered {urls}')
        for url in urls:
            Logger.log(f"url: {url}")
            for item in currency_items:
                #print(f'item: {item} url: {url}')
                #print(f'item: {type(item)} url: {type(url)}')
                #print(url)
                if str(item.lower()) in str(url[0]).lower():
                    Logger.log(f"item: {item} url: {url}")
                    filtered_urls.append(url)
                    break
            
        return filtered_urls
###############################

import threading
import requests
import random
import time
COUNT=0
class Azure_Replace():
    
    @staticmethod
    def send_request(url, brand_settings):
        global COUNT
        if not Azure_Replace.flag_javascript(Azure_Replace.extract_domain(url),brand_settings):
            payload = { 'api_key': '', 'url': url, "country_code": "us"} 
            r = requests.get('https://api.scraperapi.com/', params=payload)
            r.raise_for_status()
            Logger.log(f"Status Code:{r.status_code}")
            if r.status_code=="500":
                if COUNT<3:
                    COUNT+=1
                    Azure_Replace.send_request(url, brand_settings)
            elif r.status_code=="404" and "google" not in url:
                cached_url = f"https://webcache.googleusercontent.com/search?q=cache:{url}"
                Azure_Replace.send_request(cached_url, brand_settings)
            COUNT=0
            return r.text if r else ""
        else:
            payload = { 'api_key': '', 'url': url, "render": True, "country_code": "us"}
            r = requests.get('https://api.scraperapi.com/', params=payload)
            r.raise_for_status()
            Logger.log(f"Status Code:{r.status_code}")
            if r.status_code=="500":
                if COUNT<3:
                    COUNT+=1
                    Azure_Replace.send_request(url, brand_settings)
            elif r.status_code=="404" and "google" not in url:
                cached_url = f"https://webcache.googleusercontent.com/search?q=cache:{url}"
                Azure_Replace.send_request(cached_url, brand_settings)
            
            COUNT=0
            return r.text if r else ""

    
    @staticmethod
    def extract_domain(url):
        parsed_url = urlparse(url)
        return parsed_url.netloc if not parsed_url.netloc.startswith("www.") else parsed_url.netloc[4:]
    
    @staticmethod
    def flag_javascript(domain,brand_settings):
        if domain in brand_settings.get("domain_hierarchy", []):
            return brand_settings.get("render", False)
        else:
            return False


class ExcelProcessor():
    
    def __init__(self,filepath, searchcol,brandcol,destcol, preprocessing_option="append", min_row=6):
        self.filepath = filepath
        self.min_row = min_row
        self.searchcol = searchcol
        self.brandcol = brandcol
        self.destcol = destcol
        self.preprocessing_option = preprocessing_option
        self.wb=load_workbook(self.filepath)
        self.ws = self.wb.active
        self.total_rows=self.calculate_rows()
        #self.images=self.get_images()
        #Create new columns
        self.make_new_col("Source URL")
        self.make_new_col("Source Type")
        self.make_new_col("Seller")
        self.make_new_col("Variation Used")
        self.finalize_changes()
    
    #def find_searchcol(self):
    #    for cell in self.ws.iter_rows(values_only=True):
    #        for index, value in enumerate(cell):
    #            if value.lower in ["sku","style", ""]:
    #                return index + 1
                
    def calculate_rows(self):
        if self.preprocessing_option=="append":
            empty_rows = 0
            for row in self.ws.iter_rows(min_row=self.min_row, values_only=True):
                # Assuming destcol is 0-based index
                if not row[self.destcol]:
                    empty_rows += 1
            return empty_rows
        else:
            return self.ws.max_row-self.min_row
    
    def read_excel(self):
        excelValues=[]
        # Iterate through the rows, considering only those without a corresponding value in destcol
        for index,row in enumerate(self.ws.iter_rows(min_row=self.min_row, values_only=True)):
            search_value=row[self.searchcol]
            dest_value = row[self.destcol]
            brand_value=row[self.brandcol]
            if (search_value is not None) and (dest_value is None or dest_value == '') and (self.preprocessing_option == "append"):
                excelValues.append({'sku': search_value, 'brand': brand_value, 'excel_row_number':index+self.min_row+1})
            elif(self.preprocessing_option != "append"):
                excelValues.append({'sku': search_value, 'brand': brand_value,'excel_row_number':index+self.min_row+1})    
        return excelValues
    
    #def get_images(self):
    #    images = {}
    #    c2e = cm_to_EMU  # Convert cm to EMUs
    #    for image in self.ws._images:
    #        # Extract and store necessary details for each image
    #        col = image.anchor._from.col
    #        row = image.anchor._from.row
    #        col_offset = c2e(image.anchor._from.colOff / 914400)  # Convert offset from EMUs to cm
    #        row_offset = c2e(image.anchor._from.rowOff / 914400)  # Convert offset from EMUs to cm
    #        anchor_str = f"{col},{row},{col_offset},{row_offset}"
    #        image_stream = io.BytesIO()
    #        pil_img = PILImage.open(image.ref)
    #        pil_img.save(image_stream, format='PNG')
    #        image_stream.seek(0)
    #        images[anchor_str] = image_stream
    #    return images


    def write_excel(self, output):
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        for product_info in output:
            row_number = int(product_info[0])-1
            self.ws.cell(row=row_number, column=self.destcol + 1).value = product_info[1]
            self.ws.cell(row=row_number, column=self.find_col_index("Source URL")).value = product_info[2]
            self.ws.cell(row=row_number, column=self.find_col_index("Source Type")).value = product_info[3]
            self.ws.cell(row=row_number, column=self.find_col_index("Seller")).value = product_info[4]
            #self.ws.cell(row=row_number, column=self.find_col_index("Variation Used")).value = product_info[5]
            self.ws.cell(row=row_number, column=self.find_col_index("Variation Used")).value = datetime.datetime.now()
            if product_info[3]=="farfetch":
                for col in range(1, self.ws.max_column + 1):
                    self.ws.cell(row=row_number, column=col).fill = orange_fill
            #p2e = pixels_to_EMU  # Convert pixels to EMUs
            #for anchor_str, image_stream in self.images.items():
            #    col, row, col_offset, row_offset = map(lambda x: float(x), anchor_str.split(','))
            #    col_offset = p2e(col_offset)  # Convert offset from cm to EMUs
            #    row_offset = p2e(row_offset)  # Convert offset from cm to EMUs
            #    h, w = pil_img.height, pil_img.width
            #    size = XDRPositiveSize2D(p2e(w), p2e(h))  # Convert size from pixels to EMUs
            #    marker = AnchorMarker(col=col, colOff=col_offset, row=row, rowOff=row_offset)
            #    anchor = OneCellAnchor(_from=marker, ext=size)
            #    openpyxl_img = OpenpyxlImage(image_stream)
            #    openpyxl_img.anchor = anchor
            #    self.ws.add_image(openpyxl_img)
        self.finalize_changes()
       
    def finalize_changes(self):
        # This method will be used to save all changes at once
        self.wb.save(self.filepath)
    
    def make_new_col(self, col_name):
        header_row = self.ws[self.min_row-1]
        if col_name not in [cell.value for cell in header_row]:
            max_col=self.ws.max_column
            self.ws.insert_cols(max_col+1)
            self.ws.cell(row=self.min_row-1, column=max_col+1).value=col_name
            
            
    def find_col_index(self, col_name):
        for cell in self.ws.iter_rows(min_row=self.min_row-1, max_row=self.min_row-1, values_only=True):
            for index, value in enumerate(cell):
                if value == col_name:
                    return index + 1 # Adjusted index to 0-based
        return None

class ProductSchema:
    def __init__(self,product_schemas, source):
        self.product_schemas = product_schemas
        self.source=source
        self.parsed_products = self.parse_product_schemas(self.product_schemas)
        
        
    def get_parsed_products(self):
        return self.parsed_products

    def parse_product_schemas(self,product_schemas):
        parsed_products = []

        for schema in product_schemas:
            if schema.get('@type') == 'Product':
                offers_info = self.extract_offers(schema)
                for offer in offers_info:
                    
                    if(offer.get('@type') == 'Offer'):
                        prices=self.get_prices(offer)
                        currency=self.get_currency(offer)
                        seller=self.get_seller(offer)
                        description=self.get_description(offer)
                        title=self.get_title(offer)
                        images=self.get_images(offer)
                        url=self.get_url(offer)
                        product_details = self.create_product_details(title,images,prices,currency,url,description,seller,schema)
                        parsed_products.append(product_details)
                        
                    elif(offer.get('@type') == 'AggregateOffer'):
                        for suboffer in self.extract_offers(offer):
                            prices=self.get_prices(suboffer)
                            currency=self.get_currency(suboffer)
                            seller=self.get_seller(suboffer)
                            description=self.get_description(suboffer)
                            title=self.get_title(suboffer)
                            images=self.get_images(suboffer)
                            url=self.get_url(suboffer)
                            product_details = self.create_product_details(title,images,prices,currency,url,description,seller,schema)
                            parsed_products.append(product_details)
        return parsed_products



    def get_title(self, data):
        if isinstance(data, dict):
            for key, value in data.items():
                if key.lower() not in ['seller','brand']:
                    if key == 'name':
                        return value
                    else:
                        result = self.get_title(value)
                        if result:
                            return result
        else: return None        
            
    def get_images(self,data):
        images = []
        if isinstance(data, dict):
            for key, value in data.items():
                if key == 'image' and isinstance(value, (list, str)):
                    if isinstance(value, list):
                        images.extend(value)
                    else:
                        images.append(value)
                else:
                    images.extend(self.get_images(value))
        elif isinstance(data, list):
            for item in data:
                images.extend(self.get_images(item))
        return images

    def get_prices(self,data):
        prices = []
        if isinstance(data, dict):
            for key, value in data.items():
                if key.lower() in ['price', 'lowprice', 'highprice'] and isinstance(value, str):
                    prices.append(value.replace("$", "").replace(",", "").replace(" ", ""))
                elif key.lower() in ['price', 'lowprice', 'highprice'] and isinstance(value, (int, float)):
                    prices.append(value)
                else:
                    prices.extend(self.get_prices(value))
        elif isinstance(data, list):
            for item in data:
                prices.extend(self.get_prices(item))
        return prices 

    def get_currency(self,data):
        if isinstance(data, dict):
            currency = data.get('priceCurrency', None)
            if currency:
                return currency
            for value in data.values():
                result = self.get_currency(value)
                if result:
                    return result
        elif isinstance(data, list):
            for item in data:
                result = self.get_currency(item)
                if result:
                    return result
    def get_url(self,data):
        if self.source=="modesens":
            if isinstance(data, dict):
                url = data.get('url', None)
                if url:
                    return f"https://modesens.com{url}"
                for value in data.values():
                    result = self.get_url(value)
                    if result:
                        return f"https://modesens.com{url}"
            elif isinstance(data, list):
                for item in data:
                    result = self.get_url(item)
                    if result:
                        return f"https://modesens.com{url}"
        else:
            if isinstance(data, dict):
                url = data.get('url', None)
                if url:
                    return f"{url}"
                for value in data.values():
                    result = self.get_url(value)
                    if result:
                        return f"{url}"
            elif isinstance(data, list):
                for item in data:
                    result = self.get_url(item)
                    if result:
                        return f"{url}"
        
                
                
    def get_description(self,data):
        if isinstance(data, dict):
            for key, value in data.items():
                if key == 'description':
                    return value
                else:
                    result = self.get_description(value)
                    if result:
                        return result
                    
    def get_seller(self,data):
        if isinstance(data, dict):
            seller = data.get('seller', None)
            if isinstance(seller, dict) and 'name' in seller:
                return seller['name']
            for value in data.values():
                result = self.get_seller(value)
                if result:
                    return result
        elif isinstance(data, list):
            for item in data:
                result = self.get_seller(item)
                if result:
                    return result


    def extract_offers(self,data):
        offers = []

        if isinstance(data, dict):
            if 'offers' in data:
                # Directly append the offer or aggregate offer object
                offer_data = data['offers']
                if isinstance(offer_data, list):
                    offers.extend(offer_data)  # List of individual offers
                else:
                    offers.append(offer_data)  # Single or aggregate offer
            else:
                # Recursively search for offers in other dictionary values
                for value in data.values():
                    offers.extend(self.extract_offers(value))

        elif isinstance(data, list):
            # If the data is a list, apply the function to each element
            for item in data:
                offers.extend(self.extract_offers(item))

        return offers

    def create_product_details(self, title,images,prices,currency,url,description,seller,schema):
        product_details = {
                        'title': title,  
                        'images': images,  
                        'prices': prices,
                        'currency': currency,
                        'url': url,  
                        'description': description,  
                        'seller': seller.lower() if seller else None
                    }
        for key, value in product_details.items():
            if value in [None,[],"",{}]:
                if key == 'title':
                    product_details[key] = self.get_title(schema)
                elif key == 'images':
                    product_details[key] = self.get_images(schema)
                elif key == 'prices':
                    product_details[key] = self.get_prices(schema)
                elif key == 'currency':
                    product_details[key] = self.get_currency(schema)
                elif key == 'url':
                    product_details[key] = self.get_url(schema)
                elif key == 'description':
                    product_details[key] = self.get_description(schema)
                elif key == 'seller':
                    seller = self.get_seller(schema)
                    product_details[key] = seller.lower() if seller else seller
        return product_details
    
    
 
class Logger():
    def __init__(self, file_name):
        self.input_file_name=file_name
        self.logger=self.setup_logger()
    
    
    
    def setup_logger(self):
        logger = logging.getLogger('DataLogger')
        logger.setLevel(logging.INFO)  # Set the logging level

        # Create file handler which logs even debug messages
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        log_file_path = f"msrp_app/log_storage/{self.input_file_name}_{timestamp}_data_log.log"
        
        fh = logging.FileHandler(log_file_path, encoding='utf-8')
        fh.setLevel(logging.INFO)

        # Create formatter and add it to the handlers
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)

        # Add the handlers to the logger
        logger.addHandler(fh)

        return logger
    
    @staticmethod
    def log_product(product):
        dict={}
        logger = logging.getLogger('DataLogger')
        dict['sku']=product.input_sku
        dict['source_type']=product.source_type
        dict['currency']=product.currency
        dict['prices']=product.prices
        dict['url']=product.url
        dict['title']=product.title
        dict['images']=product.images
        dict['description']=product.description
        dict['seller']=product.seller
        dict['excel_row_number']=product.excel_row_number
        logger.info(f"{product.input_sku} Product: {dict}")
        
    @staticmethod
    def log(message):
        logger = logging.getLogger('DataLogger')
        logger.info(message)



#class FarfetchParser():
    #def __init__(self, url, html):
    #    self.url = url
    #    self.html=html
    #    self.price=self.fetch_price()
#
    #def fetch_price(self):
    #    soup = BeautifulSoup(self.html, 'html.parser')
    #    price_tag = soup.find("p", {"data-component": "PriceOriginal", "class": "ltr-vgsor4-Footnote"})
    #    price_meta_tag = soup.find("meta", property="twitter:data1")
    #    print(price_tag)
    #    if price_tag and price_tag.get_text().strip()!="Â":
    #        return list(price_tag.get_text().strip())
    #    elif price_meta_tag and 'content' in price_meta_tag.attrs:
    #        return list(price_meta_tag['content'])
    #    else:
    #        return "Price not found"
    
class ModesensParser():
    def __init__(self, html):
        self.soup=BeautifulSoup(html, 'html.parser')
        self.blocks=self.extract_blocks()
        self.product_details=self.get_product_details()
        
    
    def extract_blocks(self):
        blocks = self.soup.find_all('div', class_='d-inline-block each-list-con')
        return blocks
    
    def get_product_details(self):
        product_details = [] 

        for block in self.blocks:
            # Handle different types of price blocks
            product_detail={}
            price_box = block.find('div', class_='price-box') or block.find('span', class_='price-box')
            merchant_name = block.find('div', class_='merchant-name')
            
            # Extracting seller
            seller = merchant_name.get_text(strip=True) if merchant_name else None
            prices = []
            if price_box:
                # Find all span elements that potentially contain prices
                price_spans = price_box.find_all('span', class_='position-relative') or [price_box]
                for span in price_spans:
                    # Extracting numeric part of the price
                    price_text = span.get_text(strip=True)
                    match = re.search(r'\d+(?:\.\d+)?', price_text)
                    
                    if match:
                        price = float(match.group())
                        prices.append(price)

                    # Extracting currency symbol
                    currency = price_text[0] if price_text else None

            # Store the highest price, seller, and currency
            if prices:
                highest_price = max(prices)
                product_detail['price']=highest_price
                product_detail['seller']=seller
                product_detail['currency']=currency
                product_details.append(product_detail)

        return product_details
    
# settings = json.loads(open('msrp_project\msrp_app\settings.json').read())
# brand_settings = BrandSettings(settings)
# brand_rules = brand_settings.get_rules_for_brand("Dolce & Gabbana")
# html=Azure_Replace.send_request("https://modesens.com/product/bottega-veneta-loop-mini-intrecciato-crossbody-bag-red-61320856/", brand_rules)
# modesens=ModesensParser(html)
# print(modesens.product_details)
# with open('GGGGGGG.html', 'w',encoding='utf-8') as file:
#    file.write(html)
